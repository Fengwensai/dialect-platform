"""Excel 词表解析。

默认列头：编号 / 方言点 / 词条内容 / 例句 / 备注（可选 发音提示）。
表头与默认不一致时自动映射，未识别列由前端手动映射到目标字段。
"""
from io import BytesIO

from openpyxl import load_workbook

# 目标字段（入库字段）
FIELDS = [
    "code",
    "dialect_point",
    "content",
    "example_sentence",
    "remark",
    "pronunciation_hint",
]

FIELD_LABELS = {
    "code": "编号",
    "dialect_point": "方言点",
    "content": "词条内容",
    "example_sentence": "例句",
    "remark": "备注",
    "pronunciation_hint": "发音提示",
}

# 各字段可识别的表头别名
COLUMN_ALIASES: dict[str, list[str]] = {
    "code": ["编号", "词条编号", "序号", "id", "word_id"],
    "dialect_point": ["方言点", "方言点名称", "方言区", "发音点", "地区", "point"],
    "content": ["词条内容", "词条", "词语", "方言词", "内容", "word", "content"],
    "example_sentence": ["例句", "示例句", "例句内容", "sentence"],
    "remark": ["备注", "注释", "说明", "remark"],
    "pronunciation_hint": ["发音提示", "发音", "拼音", "注音", "同音字", "hint"],
}


def _normalize_header(h) -> str:
    return (h or "").strip().lower()


def auto_map(headers: list[str]) -> dict[str, str]:
    """表头(按列索引) -> 目标字段。key 为列索引字符串，前端据此下拉映射。"""
    alias_index: dict[str, str] = {}
    for field, aliases in COLUMN_ALIASES.items():
        for a in aliases:
            alias_index.setdefault(_normalize_header(a), field)

    mapping: dict[str, str] = {}
    for i, h in enumerate(headers):
        field = alias_index.get(_normalize_header(h))
        if field is not None:
            mapping[str(i)] = field
    return mapping


def parse_workbook(
    file_bytes: bytes, filename: str | None = None
) -> tuple[str, list[str], dict[str, str], list[dict]]:
    """解析首个工作表。

    返回 (sheet_name, headers, mapping, rows, raw_rows)：
    rows 为已按目标字段整理的 dict 列表，raw_rows 为每行的原始单元格字符串，
    两者位置一一对应，仅包含"词条内容"非空的有效行。
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    headers: list[str] = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        v = cell.value
        headers.append("" if v is None else str(v).strip())

    mapping = auto_map(headers)
    field_columns = {int(k) for k in mapping.keys()}

    rows: list[dict] = []
    raw_rows: list[list[str]] = []
    for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        item: dict = {f: "" for f in FIELDS}
        for idx in field_columns:
            if idx < len(row) and mapping.get(str(idx)):
                v = row[idx]
                item[mapping[str(idx)]] = "" if v is None else str(v).strip()
        if not item.get("content"):
            continue
        item["_row"] = excel_row_no
        rows.append(item)
        raw_rows.append(["" if v is None else str(v).strip() for v in row])

    wb.close()
    return ws.title, headers, mapping, rows, raw_rows
